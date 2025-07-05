import argparse
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# La fonction charge le fichier file_path et extrait les données de la feuille sheet_name
# en affichant les valeurs des cellules ainsi que les styles de police (taille et couleur).
def parse_excel(file_path, sheet_name):
    # Charger le fichier Excel avec openpyxl
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    script_dir = os.path.dirname(os.path.abspath(__file__))

    file_res = os.path.join(script_dir, 'resultat.txt')
    file = open(file_res, 'w')

    # Afficher les données et les styles de cellule
    print("Données et styles extraits :")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                font = cell.font
                print(f"Cellule {cell.coordinate}: Valeur = {cell.value}, Taille de police = {font.size}, Couleur = {font.color}")
                file.write(f"Cellule {cell.coordinate}: Valeur = {cell.value}, Taille de police = {font.size}, Couleur = {font.color}\n")
    file.close()

    # Fermer le workbook
    wb.close()

def main():
    #arg parse flags
    parser = argparse.ArgumentParser(
        prog='parse_excel',
        description='Parse an Excel file and extract data from a specific sheet',
      )
    parser.add_argument('file_path', type=str, help='Path to the Excel file')
    parser.add_argument('sheet_name', type=str, help='Name of the sheet to parse')

    parse_excel(file_path, sheet_name)

if __name__ == "__main__":
    main()
